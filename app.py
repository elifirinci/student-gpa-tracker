import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook


class StudentGPAApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Halic University")
        self.create_widgets()

    def create_widgets(self):
        self.file_label = tk.Label(self.root, text="Open File:")
        self.file_label.grid(row=0, column=0, padx=10, pady=5)

        self.browse_button = tk.Button(self.root, text="Browse", command=self.load_file)
        self.browse_button.grid(row=0, column=1, padx=10, pady=5)

        self.id_label = tk.Label(self.root, text="Student ID:")
        self.id_label.grid(row=1, column=0, padx=10, pady=5)

        self.id_entry = tk.Entry(self.root)
        self.id_entry.grid(row=1, column=1, padx=10, pady=5)

        self.display_button = tk.Button(self.root, text="Display", command=self.display_info)
        self.display_button.grid(row=2, column=0, padx=10, pady=5)

        self.clear_button = tk.Button(self.root, text="Clear", command=self.clear_info)
        self.clear_button.grid(row=2, column=1, padx=10, pady=5)

        # Labels for displaying student information
        self.name_label = tk.Label(self.root, text="Name: ")
        self.name_label.grid(row=3, column=0, padx=10, pady=5, sticky="e")

        self.surname_label = tk.Label(self.root, text="Surname: ")
        self.surname_label.grid(row=4, column=0, padx=10, pady=5, sticky="e")

        self.gpa_label = tk.Label(self.root, text="GPA: ")
        self.gpa_label.grid(row=5, column=0, padx=10, pady=5, sticky="e")

        self.rank_label = tk.Label(self.root, text="Rank: ")
        self.rank_label.grid(row=6, column=0, padx=10, pady=5, sticky="e")

        self.name_var = tk.StringVar()
        self.surname_var = tk.StringVar()
        self.gpa_var = tk.StringVar()
        self.rank_var = tk.StringVar()

        self.name_display = tk.Label(self.root, textvariable=self.name_var)
        self.name_display.grid(row=3, column=1, padx=10, pady=5, sticky="w")

        self.surname_display = tk.Label(self.root, textvariable=self.surname_var)
        self.surname_display.grid(row=4, column=1, padx=10, pady=5, sticky="w")

        self.gpa_display = tk.Label(self.root, textvariable=self.gpa_var)
        self.gpa_display.grid(row=5, column=1, padx=10, pady=5, sticky="w")

        self.rank_display = tk.Label(self.root, textvariable=self.rank_var)
        self.rank_display.grid(row=6, column=1, padx=10, pady=5, sticky="w")

        self.file_type_label = tk.Label(self.root, text="Save as:")
        self.file_type_label.grid(row=7, column=0, padx=10, pady=5)

        self.file_type_var = tk.StringVar(value="txt")
        self.file_type_menu = tk.OptionMenu(self.root, self.file_type_var, "xls", "txt")
        self.file_type_menu.grid(row=7, column=1, padx=10, pady=5)

        self.export_button = tk.Button(self.root, text="Export as File", command=self.export_info)
        self.export_button.grid(row=8, columnspan=2, padx=10, pady=5)

    def load_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.wb = load_workbook(file_path)
            self.ws = self.wb.active
            messagebox.showinfo("Success", "File loaded successfully.")
            self.calculate_gpa_and_rank()

    def calculate_gpa_and_rank(self):
        data = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            name, surname, student_id, physics, calculus, programming, chemistry = row
            gpa = (physics * 0.25 + calculus * 0.25 + programming * 0.30 + chemistry * 0.20)
            data.append((name, surname, student_id, gpa))
        data.sort(key=lambda x: x[3], reverse=True)
        self.data = {student[2]: student + (rank + 1,) for rank, student in enumerate(data)}

    def display_info(self):
        student_id = int(self.id_entry.get())
        if student_id in self.data:
            name, surname, student_id, gpa, rank = self.data[student_id]
            self.name_var.set(name)
            self.surname_var.set(surname)
            self.gpa_var.set(f"{gpa:.2f}")
            self.rank_var.set(rank)
        else:
            messagebox.showerror("Error", "Student ID not found.")

    def clear_info(self):
        self.name_var.set("")
        self.surname_var.set("")
        self.gpa_var.set("")
        self.rank_var.set("")
        self.id_entry.delete(0, tk.END)

    def export_info(self):
        student_info = self.result_label.cget("text")
        if student_info:
            file_type = self.file_type_var.get()
            student_id = self.id_entry.get()
            name, surname, *_ = student_info.split(",")[0].split(":")[1].strip().split()
            file_name = f"{student_id}_{name}_{surname}.{file_type}"
            with open(file_name, "w") as file:
                file.write(f"Name: {name} {surname}\nGPA: {gpa}\nRank: {rank}")
            messagebox.showinfo("Success", f"File saved as {file_name}.")
        else:
            messagebox.showerror("Error", "No student information to save.")


if __name__ == "__main__":
    root = tk.Tk()
    app = StudentGPAApp(root)
    root.mainloop()
